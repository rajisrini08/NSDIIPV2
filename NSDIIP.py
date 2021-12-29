import difflib
from importlib import reload
import os
import json
import logging
import sys
from logging import INFO, DEBUG, ERROR, info, error
from os import listdir
from os.path import isfile, join
import traceback
import numpy
from PyPDF2 import PdfFileReader, PdfFileWriter
# from numpy.lib.function_base import delete
import pyodbc as pyodbc
from PIL import Image
from pytesseract import image_to_data,Output
import pytesseract
import signal
import re
import time
from pdf2image import convert_from_path, pdfinfo_from_path, convert_from_bytes
from glob import glob
from os import path
import numpy as np
import cv2
from dateutil import parser
import requests
from fpdf import FPDF
#from pytesseract import *
# from suffix_trees import STree
import zipfile
from zipfile import ZipFile
#from werkzeug.utils import secure_filename
from fastapi import FastAPI, File, UploadFile
from starlette.requests import Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from typing import List, Tuple
import uvicorn
import shutil
import io
import csv
import boto3
import docx
from datetime import datetime, timedelta
from botocore.exceptions import NoCredentialsError
from openpyxl import load_workbook
import zipfile
import PreprocessConfig
import re
from urllib.parse import quote
import pprint
from invoice2data import extract_data
from invoice2data.extract.loader import read_templates
from subprocess import call
from pikepdf import Pdf
from PreprocessConfig import tmpdirmerge,MergeDirS3

LOGFILENAME = "NSD_IIP.log"
FORMAT = '%(asctime)-15s %(message)s'
logging.basicConfig(format=FORMAT, level=INFO, filename=LOGFILENAME)
appLogger = logging.getLogger(__name__)
appLogger.setLevel(INFO)

Image.MAX_IMAGE_PIXELS = None

app = FastAPI(title="NSD Invoice Preprocessing")
app.add_middleware(
    CORSMiddleware,
    allow_origins=['*'],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

MAX_CONTENT_LENGTH = 16 * 1024 * 1024

drivers = [item for item in pyodbc.drivers()]
driver = 'ODBC Driver 17 for SQL Server'
server = PreprocessConfig.server
database = PreprocessConfig.database

ocr_database = PreprocessConfig.OCRdatabase

uid = PreprocessConfig.uid
pwd = PreprocessConfig.pwd
bucketName = PreprocessConfig.bucketName
DirS3 = PreprocessConfig.DirS3
minGoodQuality = PreprocessConfig.minGoodQuality
minAverageQuality = PreprocessConfig.minAverageQuality

junk_height_threshold = PreprocessConfig.junkHeightThreshold
Doc_Batch_Process=[]
Doc_Batch_Process_tt=[]
ocr_con_string = f'DRIVER={driver};SERVER={server};DATABASE={ocr_database};UID={uid};PWD={pwd}'

con_string = f'DRIVER={driver};SERVER={server};DATABASE={database};UID={uid};PWD={pwd}'
try:
    cnxn = pyodbc.connect(con_string)
    appLogger.info("DB Connection success")
except:
    resp = {'message': 'Database Connection Failure'}
    appLogger.error("Database Connection Failure")
    raise Exception(resp)

ALLOWED_EXTENSIONS = set(['pdf', 'PDF', 'txt', 'TXT', 'DOCX', 'docx', 'XLSX', 'xlsx', 'JPG', 'jpg', 'JPEG', 'jpeg','tif','TIF'])


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def uploadFileAWS(localFile, fileKey):
    s3 = boto3.client('s3')
    try:
        s3.upload_file(localFile, bucketName, fileKey)
        appLogger.info("File successfully uploaded to S3 bucket: " + bucketName)
        return True
    except FileNotFoundError:
        e = sys.exc_info()
        appLogger.info("Failed uploaded to S3 bucket - " + str(e))
        appLogger.info("The file was not found: " + str(e))
        return False
    except NoCredentialsError:
        e = sys.exc_info()
        appLogger.info("Failed uploaded to S3 bucket - " + str(e))
        appLogger.info("Credentials not available: " + str(e))
        return False
    except:
        e = sys.exc_info()
        appLogger.info("Failed uploaded to S3 bucket - " + str(e))
        appLogger.info("Error occured while uploading file: " + str(e))
        return False

def tiffupdateDocLocation(Batch_Id, tiffLocation):
    try:
        cnxn = pyodbc.connect(con_string)
        appLogger.info("DB Connection success")
        cursor = cnxn.cursor()
        query = "UPDATE " + database + ".dbo.Documents SET " + database + ".dbo.Documents.Tif_file_location='" + tiffLocation + "' where " + database + ".dbo.Documents.Document_Id=" + Batch_Id + ";"
        print(query)
        cursor.execute(query)
        cnxn.commit()
        appLogger.info("Tiff File Location Update Success in Documents Table ")
        return True
    except:
        e = sys.exc_info()
        resp = {'message': 'Database Update Failure'}
        appLogger.error("Database Update in Documents Table Failure " + str(e))
        raise Exception(resp)

def updateDocLocation(Batch_Id,UpdatedPageCount, Location, isEdited):
    try:
        cnxn = pyodbc.connect(con_string)
        appLogger.info("DB Connection success")
        cursor = cnxn.cursor()
        if isEdited:
            query = "UPDATE " + database + ".dbo.Documents SET " + database + ".dbo.Documents.UpdatedFile_Location='" + Location + "',isEdited=0 where " + database + ".dbo.Documents.Document_Id=" + Batch_Id + ";"

        else:

            query = "UPDATE " + database + ".dbo.Documents SET " + database + ".dbo.Documents.Storage_Location='" + Location + "',isEdited=1, UpdatedFile_Page_Count="+str(UpdatedPageCount)+"  where " + database + ".dbo.Documents.Document_Id=" + Batch_Id + ";"
        cursor.execute(query)
        cnxn.commit()
        appLogger.info("File Location Update Success in Documents Table ")
        return True
    except:
        e = sys.exc_info()
        resp = {'message': 'Database Update Failure'}
        appLogger.error("Database Update in Documents Table Failure " + str(e))
        raise Exception(resp)


def downloadFileObjAWS(fileKey):
    try:
        s3 = boto3.resource('s3')
        bucket = s3.Bucket(bucketName)
        for files in bucket.objects.filter(Prefix=fileKey):
            filesKey = files.key
            if "/Edited/" not in filesKey:
                bucket.download_file(filesKey, os.path.basename(filesKey))
        return True
    except FileNotFoundError:
        e = sys.exc_info()
        appLogger.exception("The file was not found: " + str(e))
        return False
    except NoCredentialsError:
        e = sys.exc_info()
        appLogger.exception("Credentials not available: " + str(e))
        return False
    except:
        e = sys.exc_info()
        appLogger.exception("Error occured while uploading file: " + str(e))
        return False


def isExistsAWS(FolderPath):
    try:
        s3 = boto3.resource('s3')
        bucket = s3.Bucket(bucketName)
        bucket.objects.filter(Prefix=FolderPath)
        for object_summary in bucket.objects.filter(Prefix=FolderPath):
            return True
        return False

    except:
        e = sys.exc_info()
        appLogger.info("Failed to verify S3 bucket - " + str(e))
        appLogger.exception("Error occured while verifying S3 folder exists: " + str(e))
        return False


def insertBatchDetail(fileName, pgCount, batchName):
    try:
        cnxn = pyodbc.connect(con_string)
        appLogger.info(" has DB Connection success")
        cursor = cnxn.cursor()
        cursor.execute(
            "Insert into " + database + ".dbo.Doc_Batch(Doc_Name,No_of_Pages,Batch_Name) values('" + fileName + "'," + str(
                pgCount) + ",'" + batchName + "');")
        cnxn.commit()
        cursor.execute("Select SCOPE_IDENTITY();")
        batchID = str(cursor.fetchone()[0])
        appLogger.info("Insert Batch detail in Doc_Batch Success")
        return batchID
    except:
        e = sys.exc_info()
        resp = {'status': 501, 'message': 'Database Update Failure '}
        appLogger.error("Database Insert Failure " + str(e))
        raise Exception(resp)


def getData(tableName):
    try:
        cnxn = pyodbc.connect(con_string)
        cursor = cnxn.cursor()
        cursor.execute("SELECT * from " + database + ".dbo." + tableName)
        rows = cursor.fetchall()
        if len(rows) == 0:
            resp = {'status': 501, 'message': 'No data from database'}
            appLogger.error("No data from database")
            raise Exception(resp)
        cnxn.close()
        appLogger.info("Insert Batch detail in Doc_Batch Success")
        return rows
    except:
        e = sys.exc_info()
        resp = {'status': 501, 'message': 'Database Update Failure '}
        appLogger.error("Database Fetch Data Failure " + str(e))
        raise Exception(resp)

"""BatchProcessID,BatchID,KeyID,Occurance"""
"""Doc_Batch_Process: [(108375, 20668, 1, 2), (108376, 20668, 2, 2), (108377, 20668, 3, 4), 
                        (108378, 20668, 9, 1), (108379, 20668, 10, 1), (108380, 20668, 11, 1), 
                        (108381, 20668, 12, 1), (108382, 20668, 19, 1), (108383, 20668, 20, 16), 
                        (108384, 20668, 31, 3), (108385, 20668, 32, 4)]"""
def insertBatchProcess(batchID, keyID, Occurance):
    try:
        cnxn = pyodbc.connect(con_string)
        appLogger.info("DB Connection success")
        cursor = cnxn.cursor()
        cursor.execute("Insert into " + database + ".dbo.Doc_Batch_Process(Batch_Id,Key_Id,Occurance) values(" + str(
            batchID) + "," + str(keyID) + "," + str(Occurance) + ");")
        cnxn.commit()
        cursor.execute("Select SCOPE_IDENTITY();")
        batchProcessID = str(cursor.fetchone()[0])
        appLogger.info("File Location Update Success in Documents Batch Process Table ")
        return batchProcessID
    except:
        e = sys.exc_info()
        resp = {'message': 'Database Update Failure'}
        appLogger.error("Database Update Failure" + str(e))
        raise Exception(resp)


def isExistsAWS(FolderPath):
    try:
        s3 = boto3.resource('s3')
        bucket = s3.Bucket(bucketName)
        for object_summary in bucket.objects.filter(Prefix=FolderPath):
            return True
        return False

    except:
        e = sys.exc_info()
        appLogger.info("Failed to verify S3 bucket - " + str(e))
        appLogger.exception("Error occured while verifying S3 folder exists: " + str(e))
        return False

"""BatchProcessID,BatchID,KeyID,Occurance"""
"""from SQL Doc_Batch_Process: [(108416, 20675, 2, 2), (108417, 20675, 3, 11), (108418, 20675, 5, 1), 
                        (108419, 20675, 7, 1), (108420, 20675, 10, 1), (108421, 20675, 20, 1)]"""

"""from Local Doc_Batch_Process: [(100000, 20000, 2, 2), (100000, 20000, 3, 11), (100000, 20000, 5, 1),
                                  (100000, 20000, 7, 1), (100000, 20000, 10, 1), (100000, 20000, 20, 1)]"""
def getDocType(Batch_Id, keyMapping, docMaster):
# def getDocType(keyMapping, docMaster):
    try:
        cursor = cnxn.cursor()
        cursor.execute("SELECT * from " + database + ".dbo.Doc_Batch_Process where Batch_Id='" + str(Batch_Id) + "';")
        Doc_Batch_Process = cursor.fetchall()
        ##print("Doc_Batch_Process:",Doc_Batch_Process)
        if len(Doc_Batch_Process) == 0:
            appLogger.error("No data from Doc_Batch_Process table, returning 'Unknown'")
            return 'Unknown', 'N/A', 'N/A'

        docMasterKeyId = []
        docMasterDocType = []
        npKeyMapping = np.array(keyMapping)
        resultDict = {}

        for i in range(0, len(docMaster)):
            docMasterKeyId.append(docMaster[i][0])
            docMasterDocType.append(docMaster[i][1])
        for row in range(0, len(Doc_Batch_Process)):
            selectRows = np.where(npKeyMapping[:, 2] == Doc_Batch_Process[row][2])
            result = npKeyMapping[selectRows]
            if result[0][1] in resultDict:
                val = resultDict[result[0][1]] + result[0][4]
            else:
                val = result[0][4]
            resultDict.update({result[0][1]: val})

        dictValues = list(resultDict.values())
        maxValue = dictValues.index(max(dictValues))
        docTypeKey = docMasterKeyId.index(list(resultDict.keys())[maxValue])
        docType = docMasterDocType[docTypeKey]
        percentScore = (max(dictValues) / sum(dictValues)) * 100

        ##print("dictValues:",dictValues)
        if max(dictValues) > 10.0:
            return docType, max(dictValues), round(percentScore, 2)
        else:
            return 'Unknown', 'N/A', 'N/A'

    except:
        e = sys.exc_info()
        err = e[1]
        resp = {'message': 'Error occured-' + str(err)}
        appLogger.error("Exception occured in getting document type -" + str(e))
        raise Exception(resp)

# def getDocType_tt(keyMapping, docMaster):
def getDocType_tt(Batch_Id, keyMapping, docMaster):
    try:
        cursor = cnxn.cursor()
        cursor.execute("SELECT * from " + database + ".dbo.Doc_Batch_Process where Batch_Id='" + str(Batch_Id) + "';")
        Doc_Batch_Process = cursor.fetchall()
        ##print("Doc_Batch_Process_tt:",Doc_Batch_Process_tt)
        if len(Doc_Batch_Process_tt) == 0:
            appLogger.error("No data from Doc_Batch_Process table, returning 'Unknown'")
            return 'Unknown', 'N/A', 'N/A'

        docMasterKeyId = []
        docMasterDocType = []
        npKeyMapping = np.array(keyMapping)
        resultDict = {}

        for i in range(0, len(docMaster)):
            docMasterKeyId.append(docMaster[i][0])
            docMasterDocType.append(docMaster[i][1])
        for row in range(0, len(Doc_Batch_Process_tt)):
            selectRows = np.where(npKeyMapping[:, 2] == Doc_Batch_Process_tt[row][2])
            result = npKeyMapping[selectRows]
            if result[0][1] in resultDict:
                val = resultDict[result[0][1]] + result[0][4]
            else:
                val = result[0][4]
            resultDict.update({result[0][1]: val})

        dictValues = list(resultDict.values())
        maxValue = dictValues.index(max(dictValues))
        docTypeKey = docMasterKeyId.index(list(resultDict.keys())[maxValue])
        docType = docMasterDocType[docTypeKey]
        percentScore = (max(dictValues) / sum(dictValues)) * 100

        ##print("dictValues:",dictValues)
        if max(dictValues) > 10.0:
            return docType, max(dictValues), round(percentScore, 2)
        else:
            return 'Unknown', 'N/A', 'N/A'

    except:
        e = sys.exc_info()
        err = e[1]
        resp = {'message': 'Error occured-' + str(err)}
        appLogger.error("Exception occured in getting document type -" + str(e))
        raise Exception(resp)

def imageOcr(file_name):
    # OCR reading
    try:
        im = Image.open(file_name)
        ocr_output = pytesseract.image_to_string(im)
        out_text = ocr_output
        im.close()
        return out_text
    except:
        e = sys.exc_info()
        err = str(e[1])
        resp = {'status': 501, 'Exception': err}
        raise Exception(resp)


def readDocx(fileName):
    try:
        open_doc = open(fileName, 'rb')
        doc = docx.Document(open_doc)
        fullText = []
        result = [p.text for p in doc.paragraphs]
        for p in result:
            fullText.append(p)
        text = '\n'.join(fullText)
        return text
    except:
        e = sys.exc_info()
        err = str(e[1])
        resp = {'status': 501, 'Exception': err}
        raise Exception(resp)


def readXlsx(fileName):
    try:
        allinfo = []
        workbook = load_workbook(fileName)
        for n, sheet in enumerate(workbook.worksheets):
            worksheet = workbook[sheet.title]
            string_info = ''
            for row_cells in worksheet.iter_rows():
                for cell in row_cells:
                    string_info += str(cell.value) + ", "
            allinfo.append(string_info[:-2])
        return allinfo
    except:
        e = sys.exc_info()
        err = str(e[1])
        resp = {'status': 501, 'Exception': err}
        raise Exception(resp)


def docx_get_characters_number(path):
    try:
        document = zipfile.ZipFile(path)
        xml_content = document.read('docProps/app.xml')
        document.close()
        regex = "<Pages>(.+?)</Pages>"
        pattern = re.compile(regex)
        return re.findall(pattern, xml_content.decode('ascii'))[0]
    except:
        return '1'

""" 
results = image_to_data(myfile, output_type=’data.frame’)

myfile: the pre-processed image

output_type=’data.frame’: We output the data into a dataframe.
"""
def getConfidence(myfile, Doc_Id, total):
    reqTempDir = 'tmp/' + Doc_Id
    reqResDir = 'res/' + Doc_Id
    try:
        try:
            for line in os.popen("ps ax | grep tesseract | grep -v grep"):
                fields = line.split()
                pid = fields[0]
                os.kill(int(pid), signal.SIGKILL)
            results = image_to_data(myfile, output_type= Output.DICT)

        except:
           e = sys.exc_info()
           traceback.print_exc()
           if os.path.exists(reqResDir):
            #    shutil.rmtree("imag/")
                os.system("rm -rf " + reqResDir)
                os.system("rm -rf " + reqTempDir)

           appLogger.exception("Error occured while getting confidence: " + str(e))
           err = str(e[1])
           resp = {'status': 501, 'Exception': err}
           raise Exception(resp)

        total_conf = []
        text = ""

        # code to determeine if the page is junk based on key words and height
        # changes NP-------

        isJunk = ""

        junkStringList=['dellvery receipt','delivery receipt']

        # print("junkStringList:",junkStringList)

        tempr = ""
        for j in range(len(results["text"])):
            if results["height"][j] > junk_height_threshold:
                tempr = tempr + " " + results["text"][j]
        tempr = tempr.lower()


        junkstrL=[]
        for junkstr in junkStringList:

            if junkstr.lower() in tempr:
                junkstrL.append(junkstr)
                isJunk = "True"
                break
        # --------------

        for i in range(0, len(results["text"])):
            text = text + " " + results["text"][i]
            conf = int(results["conf"][i])

            # filter out weak confidence text localizations

            if conf > 0:
                total_conf.append(conf)
                text = "".join(text).strip()

                if (len(total_conf) != 0):
                    total.append(sum(total_conf) / len(total_conf))

        print("OCR Text:",text)        
        print("Temporary Text:",tempr)
        print("isJunk status for above page:",isJunk,"-->",junkstrL)
        print()
        return (isJunk,total,text,results)
    except:
        e = sys.exc_info()
        traceback.print_exc()
        appLogger.exception("Error occured while getting confidence: " + str(e))
        return False

def IdentifyVendor(OcrText):
    result = {}
    cnxn = pyodbc.connect(ocr_con_string)
    cursor = cnxn.cursor()
    query = "select VendorName ,HeaderPattern    from " + ocr_database + ".dbo.Vendors ;"
    cursor.execute(query)
    query_result = cursor.fetchall()
    result={}
    cnxn.close()

    for res in query_result:

        vendorname= res[0]
        headerOCR= res[1]
        result[vendorname] =headerOCR

    for key,value in result.items():

        similarity = difflib.SequenceMatcher(None, OcrText, value).ratio()
        if similarity > 0.6:
            return key

    return "Unknown"


def IdentifyDuplicatePages(Ocr_File__text_output):
    similarity_scores = []
    for i in range(len(Ocr_File__text_output)):

        for j in range(i + 1, len(Ocr_File__text_output)):

            similarity = difflib.SequenceMatcher(None, Ocr_File__text_output[i], Ocr_File__text_output[j]).ratio()
            if similarity > 0.95:
                pair = [i, j, similarity]
                similarity_scores.append(pair)

    # pages to delete
    pages_to_delete = []
    for pair in similarity_scores:
        # checking threshold
        if pair[2] > 0.95:
            pages_to_delete.append(pair[1])

    return pages_to_delete

def GetVendorList():
    result = []
    cnxn = pyodbc.connect(ocr_con_string)
    cursor = cnxn.cursor()
    query="select VendorName   from " + ocr_database + ".dbo.Vendors ;"
    cursor.execute(query)
    query_result = cursor.fetchall()
    for item in query_result:
        result.append(item[0])
    return result


def getValue(ocrData, left_top, right_bottom):
    ans = []
    for i in range(len(ocrData['left'])):
        if (left_top[0] <= ocrData['left'][i] <= right_bottom[0]) and (
                left_top[1] <= ocrData['top'][i] <= right_bottom[1]):
            ans.append(ocrData['text'][i])
    return ' '.join(ans).strip()


def InsertExractedResults(result,vendorname,docId,pagecount,targetSystemID):

    try:

        inv_number=result["InvoiceNumber"]
        inv_Date=parser.parse(result["InvoiceDate"])
        inv_total=result["InvoiceAmount"]
        currency=result["po_number"]
        packaging_slip_number=result["packaging_slip_number"]
        tax=result["Tax"]
        frieght=result["Frieght"]
        order_number=result["order_number"]
        delivery_number=result["delivery_number"]
        reference_number=result["reference_number"]


        # accountNumber=result["AccountNumber"]
        vendor_details=GetVendorDetails(vendorname)
        currency=ExtractCurrency(currency,inv_total)


        total_extracted = re.findall('\d*\.?\d+', inv_total)

        inv_total=total_extracted[0]

        #date = parser.parse(inv_Date)
        #print(date)


        today = datetime.now()

        date_time = today.strftime("%m_%d_%Y_%H%M%S")
        batchname="QBOCR_"+str(date_time)
        cnxn = pyodbc.connect(con_string)
        cursor = cnxn.cursor()


        #query to update document status !! need to move to bot
        doc_status_query = "UPDATE " + database + ".dbo.Documents SET " + database + ".dbo.Documents.status='Processed',[Target_system_Id]='"+targetSystemID+"', [Updated_On]='"+str(datetime.now().date())+"' where  " + database + ".dbo.Documents.Document_Id=" + str(docId) + ";"
        cursor.execute(doc_status_query);
        cnxn.commit()



        cursor.execute("select Created_by,Customer_Id,Dept_Id from [dbo].[Documents] where Document_Id=" + str(docId) + " ;")
        created_by_data=cursor.fetchone()
        if created_by_data:
            created_by = str(created_by_data[0])
            Customer_Id = str(created_by_data[1])
            Dept_Id = str(created_by_data[1])
        else:
            created_by=''
            Customer_Id =0
            Dept_Id = 0

        # targetSystemID='QBOCR'
        insert_batch_query="Insert into " + database + ".dbo.Batches(Customer_Id,Dept_Id,Created_by,Updated_by,Location,Date,Target_system_Id,Page_Count,ABBY_BATCH_NAME,Status) values("+str(Customer_Id)+","+str(Dept_Id)+",'"+created_by+ "','"+created_by+ "','','"+ str(datetime.now().date())+"','"+ targetSystemID + "',0,'" +batchname+ "','"+"Processed"+"');"


        cursor.execute(insert_batch_query);
        cnxn.commit()
        cursor.execute("Select SCOPE_IDENTITY();")


        batchID = str(cursor.fetchone()[0])
        insert_batch_doc_query="Insert into " + database + ".dbo.[Batch_Document_Mapping]([Batch_Id], [Document_Id], [Page_Count],Created_by,Updated_by) VALUES("+str(batchID)+","+str(docId)+","+str(pagecount)+",'"+created_by+"','"+created_by+"');"
        cursor.execute(insert_batch_doc_query);
        cnxn.commit()

        cnxn.close()
        cnxn_ocr = pyodbc.connect(ocr_con_string)
        cursor_ocr = cnxn_ocr.cursor()
        # print("Insert into " + ocr_database + ".dbo.Invoice(InvoiceNumber,Total,BatchName,InvoiceDate,Name1,City1,State1,Address1,Street1,ZIP,Country1,Currency) values('" + inv_number + "','"  + inv_total +  "','" +  batchname + "','"+ str(inv_Date)+"','"+vendor_details["VendorName"]+"','"+vendor_details["City"]+"','"+vendor_details["State"]+"','"+str(vendor_details["Address"])+"','"+vendor_details["Street"]+"','"+vendor_details["PostalCode"]+"','"+vendor_details["Country"]+"');")
        cursor_ocr.execute( "Insert into " + ocr_database + ".dbo.Invoice(InvoiceNumber,Total,BatchName,InvoiceDate,Name1,City1,State1,Address1,Street1,ZIP,Country1,Currency,Reference_No,Packaging_Slip_No,Tax_Amount,Order_No,Delivery_No,Fright_Amount) values('" + inv_number + "','"  + inv_total +  "','" +  batchname + "','"+ str(inv_Date)+"','"+vendor_details["VendorName"]+"','"+vendor_details["City"]+"','"+vendor_details["State"]+"','"+str(vendor_details["Address"])+"','"+vendor_details["Street"]+"','"+vendor_details["PostalCode"]+"','"+vendor_details["Country"]+"','"+currency+"','"+reference_number+"','"+packaging_slip_number+"','"+tax+"','"+order_number+"','"+delivery_number+"','"+frieght+"');")


        cursor_ocr.execute("Select SCOPE_IDENTITY();")
        invoiceRowIndex = str(cursor_ocr.fetchone()[0])
        # print(result.keys())
        # print(result['lines'])


        if 'lines' in result.keys():


            for item in result['lines']:
                lineitemfields =item.keys()
                ##print(item)
                if 'ItemCode' in lineitemfields:
                        ItemCode=item['ItemCode']
                else:
                    ItemCode=''

                if 'Description' in lineitemfields:
                        description=item['Description']
                else:
                    description=''

                if 'Quantity' in lineitemfields:
                        Quantity=item['Quantity']
                else:
                    Quantity=0

                if 'UnitPrice' in lineitemfields:
                        UnitPrice=item['UnitPrice']
                else:
                    UnitPrice=0


                if 'TotalPriceNetto' in lineitemfields:
                        TotalPriceNetto=item['TotalPriceNetto']
                else:
                    TotalPriceNetto=0

                if 'UOM' in lineitemfields:
                        UOM=item['UOM']
                else:
                    UOM='Each'

                if 'Position' in lineitemfields:
                        Position=item['Position']
                else:
                    Position=''

                ##print("Insert into " + ocr_database + ".dbo.LineItems(Invoice_ROW_INDEX,ItemCode,Quantity,UnitPrice,Description,TotalPriceNetto) values("+str(invoiceRowIndex)+",'"+ItemCode+"',"+str(Quantity)+","+str(UnitPrice)+",'"+description+"','"+TotalPriceNetto+" ')")


                cursor_ocr.execute("Insert into " + ocr_database + ".dbo.LineItems(Invoice_ROW_INDEX,ItemCode,Quantity,UnitPrice,Description,TotalPriceNetto,Position) values("+str(invoiceRowIndex)+",'"+ItemCode+"',"+str(Quantity)+","+str(UnitPrice)+",'"+description+"','"+TotalPriceNetto+"','"+Position+" ')")


        # print(result['LineItems'])
        if vendor_details['VendorType']=='Non-PO':

            description='Default LineItem for ' + str(vendorname)
            TotalPriceNetto=inv_total
                # print("Insert into " + ocr_database + ".dbo.LineItems(Invoice_ROW_INDEX,Description,TotalPriceNetto) values("+str(invoiceRowIndex)+",'"+description+"','"+TotalPriceNetto+" ')")

            cursor_ocr.execute( "Insert into " + ocr_database + ".dbo.LineItems(Invoice_ROW_INDEX,Description,TotalPriceNetto) values("+str(invoiceRowIndex)+",'"+description+"','"+TotalPriceNetto+" ')")


        # cursor_ocr.execute( "Insert into " + ocr_database + ".dbo.Lineitems(Description) values(' ')")
        cnxn_ocr.commit()
        cnxn_ocr.close()


        appLogger.info("Insert ExtractData in ocr_database Invoice")
        return batchname
    except:
        e = sys.exc_info()
        resp = {'status': 501, 'message': 'Database Update Failure '}
        raise Exception(resp)

    return


def GetVendorDetails(vendor_name):
    result = {}
    cnxn = pyodbc.connect(ocr_con_string)
    cursor = cnxn.cursor()
    query="select VendorName, State ,City ,Street,PostalCode ,Country,Address,VendorType    from " + ocr_database + ".dbo.Vendors where VendorName='" + vendor_name + "';"
    cursor.execute(query)
    query_result = cursor.fetchone()

    result["VendorName"] = query_result[0]
    result["State"] = query_result[1]
    result["City"] = query_result[2]
    result["Street"] = query_result[3]
    result["PostalCode"] = query_result[4]
    result["Country"] = query_result[5]
    result["Address"] = query_result[6]
    result["VendorType"] = query_result[7]
    # print(result)
    return result


def ValidateOCRresults(resultsToValidate):
    isValid=True
    #pass dictionary of extracted result with key as field name
    fields=resultsToValidate.keys()
    if "InvoiceNumber" not in fields or "InvoiceDate" not in fields or "InvoiceAmount" not in fields :
        isValid = False


    for key in resultsToValidate:
        if key=='InvoiceNumber':
            if resultsToValidate[key].strip()=="":
                isValid=False

            elif key == 'InvoiceDate':
                    if resultsToValidate[key].strip() == "":
                        isValid = False


            elif key == 'InvoiceAmount':
                if resultsToValidate[key].strip() == "":
                    isValid = False

    return  isValid

def GetFieldBoundings(vendorName):
    boundings=[]

    cnxn = pyodbc.connect(ocr_con_string)
    cursor = cnxn.cursor()
    query = "select FieldName  ,bb.left_,	bb.top_,bb.	width_,bb.height_ from " + ocr_database +".[dbo].[BoundingBoxCordinates] bb inner join " + ocr_database +".[dbo].[TemplateFields] tf  on" \
            " bb.FieldID= tf.[FieldID] inner join "+ ocr_database +" .[dbo].[Vendors] ve on ve.VendorId =bb.VendorId where VendorName='"+vendorName+"';"

    cursor.execute(query)
    query_result = cursor.fetchall()
    cnxn.close()
    for item in query_result:
        bound={}
        bound["FieldName"]=item[0]
        bound["left"]=item[1]
        bound["top"]=item[2]
        bound["width"]=item[3]
        bound["height"]=item[4]
        boundings.append(bound)

    return boundings


def ExtractData(OcrData, vendor):
    result = {}
    vendorBoundings=GetFieldBoundings(vendor)

    for bound in vendorBoundings:
        fieldname = bound["FieldName"]

        left_top = (bound['left'], bound['top'])
        right_bottom = (bound['left'] + bound['width'] + 50,
                        bound['top'] + bound['height'] + 50)

        testValue = getValue(OcrData, left_top, right_bottom)
        result[fieldname] = testValue
    total_extracted = re.findall('\d*\.?\d+', result['InvoiceAmount'])

    inv_total = total_extracted[0]
    result["LineItems"] = [{'Description':'Default LineItem for '+str(vendor),'TotalPriceNetto': ''+str(inv_total)}]

        #pp.pprint(OcrData)


    return result


def GetOCRPreference(vendorName):
    target_system_Id=''
    cnxn = pyodbc.connect(ocr_con_string)
    cursor = cnxn.cursor()
    query = "select Target_system_Id    from " + ocr_database + ".dbo.Vendors where  VendorName='"+vendorName+"';"
    cursor.execute(query)
    query_result = cursor.fetchone()
    cnxn.close()
    target_system_Id=query_result[0]

    return target_system_Id


def ExtractInvoice2Data(file, vendor):
    result = {}
    templates = read_templates('invtemplates/')
    output = extract_data(file, templates=templates)

    fieldsExtracted=output.keys()

    if "invoice_number" in fieldsExtracted:
        result["InvoiceNumber"]=str(output["invoice_number"])
    else:
        result["InvoiceNumber"] = ''

    if "date" in fieldsExtracted:
        result["InvoiceDate"]=str(output["date"])
    else:
        result["InvoiceDate"] = ''

    if "amount" in fieldsExtracted:
        result["InvoiceAmount"]=str(output["amount"])
    else:
        result["InvoiceAmount"] = ''

    if "currency" in fieldsExtracted:
        result["currency"]=str(output["currency"])
    else:
        result["currency"]='USD'

    if "issuer" in fieldsExtracted:
        result["VendorName"]=str(output["issuer"])

    if "po_number" in fieldsExtracted:
        result["po_number"] = output['po_number']
    else:
        result["po_number"] = ''

    if "reference_number" in fieldsExtracted:
        result["reference_number"] = output['reference_number']
    else:
        result["reference_number"] = ''

    if "delivery_number" in fieldsExtracted:
        result["delivery_number"] = output['delivery_number']
    else:
        result["delivery_number"] = ''


    if "Frieght" in fieldsExtracted:
        result["Frieght"] = output['Frieght']
    else:
        result["Frieght"] = ''

    if "packaging_slip_number" in fieldsExtracted:
        result["packaging_slip_number"]= str(output['packaging_slip_number'])
    else:
        result["packaging_slip_number"] = ''

    if "order_number" in fieldsExtracted:
        result["order_number"]= str(output['order_number'])
    else:
        result["order_number"] = ''


    if "Tax" in fieldsExtracted:
        result["Tax"]= str(output['Tax'])
    else:
        result["Tax"] = ''

    if "lines" in fieldsExtracted:
        result["lines"]=output["lines"]


    return result

def ExtractCurrency(currencyExtracted, InvoiceAmount):
    currency='USD'

    cnxn = pyodbc.connect(ocr_con_string)
    cursor = cnxn.cursor()
    query = "select  CurrencyPattern  , Currency    from " + ocr_database + ".dbo.CurrencyMapping ;"

    cursor.execute(query)
    query_result = cursor.fetchall()
    for item in query_result:
        if item[0] in currencyExtracted+InvoiceAmount:
            currency=item[1]
            break
    return currency

def pdf2tiff(inFileLocation):
    outFileName = os.path.splitext(inFileLocation)[0] + '.tif'
    call(['gs','-o',outFileName,'-dNOPAUSE','-sDEVICE=tiff24nc','-r300x300','-sCompression=lzw',inFileLocation])

def tiff2pdf(inFileLocation):
    outFileName = os.path.splitext(inFileLocation)[0] + '.pdf'
    call(['tiff2pdf','-o',outFileName,'-j',inFileLocation])

def removetiff(filepath):
    os.remove(filepath)

@app.post('/MergePDF')
def MergePDF(request: Request, Doc_Id: str,files: List[UploadFile] = File(...)):
    try:
        appLogger.info("Response to " + str(request.client.host) + " has Received /MergePDF request.")
        inputFilesDirAWS = os.path.join(MergeDirS3, Doc_Id)
        if not os.path.exists(tmpdirmerge):
            os.mkdir(tmpdirmerge)
        else:
            os.system("rm -rf " + tmpdirmerge)
            os.mkdir(tmpdirmerge)
        outfilename = 'Merge_' + os.path.basename(files[0].filename).replace('.zip','.pdf')

        zipfilename = files[0].filename

        for file in files:
            if file.filename != "":
                savedFileName = os.path.join(tmpdirmerge, file.filename)
                file_object = file.file
                upload_folder = open(savedFileName, 'wb+')
                shutil.copyfileobj(file_object, upload_folder)
                upload_folder.close()
        
        with ZipFile(tmpdirmerge +'/'+zipfilename, 'r') as zip: 
            zip.extractall(tmpdirmerge)
        os.system("rm -rf "+tmpdirmerge +'/'+zipfilename)

        writer = Pdf.new()
        for pdffilepath in glob(tmpdirmerge+'/'+'*.pdf'):
            src = Pdf.open(pdffilepath)
            writer.pages.extend(src.pages)

        writer.save(tmpdirmerge+'/'+outfilename)
        uploadFileAWS(tmpdirmerge+'/'+outfilename, inputFilesDirAWS + "/" + outfilename)
        fullPathAWS = "https://" + bucketName + ".s3.amazonaws.com/" + inputFilesDirAWS + "/" + quote(outfilename)
        result = {"Merge Status":"Success",
                "FilePath":fullPathAWS}
        appLogger.info(str(result))
        return result

    except Exception as e:
        print(e)
        return {"Merge Status":"Failed"}

    finally:
        os.system("rm -rf " + tmpdirmerge)

# @app.post('/MergePDF')
# def MergePDF(request: Request, Doc_Id: str,files: List[UploadFile] = File(...)):
#     try:
#         appLogger.info("Response to " + str(request.client.host) + " has Received /MergePDF request.")
#         inputFilesDirAWS = os.path.join(MergeDirS3, Doc_Id)
#         if not os.path.exists(tmpdirmerge):
#             os.mkdir(tmpdirmerge)
#         else:
#             os.system("rm -rf " + tmpdirmerge)
#             os.mkdir(tmpdirmerge)
#         outfilename = 'cat.' + os.path.basename(files[0].filename)

#         writer = Pdf.new()
#         for file in files:
#             if file.filename != "":
#                 savedFileName = os.path.join(tmpdirmerge, file.filename)
#                 file_object = file.file
#                 upload_folder = open(savedFileName, 'wb+')
#                 shutil.copyfileobj(file_object, upload_folder)
#                 upload_folder.close()
#                 src = Pdf.open(tmpdirmerge+'/'+file.filename)
#                 writer.pages.extend(src.pages)
#         writer.save(tmpdirmerge+'/'+outfilename)
#         uploadFileAWS(tmpdirmerge+'/'+outfilename, inputFilesDirAWS + "/" + outfilename)
#         fullPathAWS = "https://" + bucketName + ".s3.amazonaws.com/" + inputFilesDirAWS + "/" + quote(outfilename)
#         result = {"Merge Status":"Success",
#                 "FilePath":fullPathAWS}
#         appLogger.info(str(result))
#         return result

#     except Exception as e:
#         print(e)
#         return {"Merge Status":"Failed"}

#     finally:
#         os.system("rm -rf " + tmpdirmerge)

@app.post("/file-upload")
def upload_file(request: Request, Doc_Id: str, files: List[UploadFile] = File(...)):
    try:
        appLogger.info("Response to " + str(request.client.host) + " has Received /File-upload request.")
        reqTempDir = 'tmp/' + Doc_Id
        inputFilesDirAWS = os.path.join(DirS3, Doc_Id)
        checkAWS = isExistsAWS(inputFilesDirAWS + "/")
        if checkAWS:
            resp = {'status': 200, 'message': 'Files successfully already uploaded', 'Document_Id': Doc_Id}
            return resp
        
        if not os.path.exists(os.path.dirname(reqTempDir)):
            os.mkdir(reqTempDir)
        else:
            os.system("rm -rf " + reqTempDir)
            os.mkdir(reqTempDir)

        # fullPathAWS="https://"+bucketName+".s3.amazonaws.com/"+inputFilesDirAWS
        for file in files:
            if file.filename == '':
                appLogger.info("Response to " + str(request.client.host) + " No files Selected for uploading!!")
                resp = {'message': 'No file selected for uploading', 'status': 400}
                appLogger.info(str(request.client.host) + "No files Selected for uploading!!")
                return resp

            if file and allowed_file(file.filename):
                filename = file.filename
                appLogger.info("Response to " + str(request.client.host) + "For File name: " + str(file))
                savedFileName = os.path.join(reqTempDir, filename)
                file_object = file.file
                upload_folder = open(savedFileName, 'wb+')
                shutil.copyfileobj(file_object, upload_folder)
                upload_folder.close()

                #if input is TIF convert to PDF and upload it to AWS, delete the tif file locally
                if (os.path.splitext(filename)[1] == '.tif') or (os.path.splitext(filename)[1] == '.TIF'):
                    tiff2pdf(savedFileName)
                    pdffilename = filename.replace('tif','pdf').replace('TIF','pdf')

                    fullPathAWS = "https://" + bucketName + ".s3.amazonaws.com/" + inputFilesDirAWS + "/" + quote(filename)
                    pdffullPathAWS = "https://" + bucketName + ".s3.amazonaws.com/" + inputFilesDirAWS + "/" + quote(pdffilename)
                    uploadFileAWS(savedFileName, inputFilesDirAWS + "/" + filename)
                    uploadFileAWS(savedFileName.replace('tif','pdf').replace('TIF','pdf'), inputFilesDirAWS + "/" + pdffilename)
                    updateDocLocation(Doc_Id,0, pdffullPathAWS, False)
                    tiffupdateDocLocation(Doc_Id, fullPathAWS)
                    removetiff(reqTempDir+'/'+filename)
                    return {'status': 200, 'message': 'File successfully uploaded', 'Document_Id': Doc_Id, 'FilePath':pdffullPathAWS , 'TifFilePath':fullPathAWS}

                #elif Convert PDF to TIF and upload it to AWS, delete the tif file locally
                elif (os.path.splitext(filename)[1] == '.pdf') or (os.path.splitext(filename)[1] == '.PDF'):
                    pdf2tiff(savedFileName)
                    tiffilename = filename.replace('pdf','tif').replace('PDF','tif')

                    fullPathAWS = "https://" + bucketName + ".s3.amazonaws.com/" + inputFilesDirAWS + "/" + quote(filename)
                    tiffullPathAWS = "https://" + bucketName + ".s3.amazonaws.com/" + inputFilesDirAWS + "/" + quote(tiffilename)
                    uploadFileAWS(savedFileName, inputFilesDirAWS + "/" + filename)
                    uploadFileAWS(savedFileName.replace('pdf','tif').replace('PDF','tif'), inputFilesDirAWS + "/" + tiffilename)
                    updateDocLocation(Doc_Id,0, fullPathAWS, False)
                    tiffupdateDocLocation(Doc_Id, tiffullPathAWS)
                    removetiff(reqTempDir+'/'+tiffilename)
                    return {'status': 200, 'message': 'File successfully uploaded', 'Document_Id': Doc_Id, 'FilePath':fullPathAWS, 'TifFilePath':tiffullPathAWS}

                else:
                    fullPathAWS = "https://" + bucketName + ".s3.amazonaws.com/" + inputFilesDirAWS + "/" + quote(filename)
                    uploadFileAWS(savedFileName, inputFilesDirAWS + "/" + filename)
                    updateDocLocation(Doc_Id,0, fullPathAWS, False)


            else:
                appLogger.info("Response to " + str(request.client.host) + ": Has a Invalid file type!!")
                appLogger.info(str(request.client.host) + " : Invalid file type!!")
                appLogger.info("Received file \'" + str(file.filename) + "\' is not a PDF")
                resp = {'message': 'Allowed file type is pdf, xlsx, docx,txt,jpeg', 'status': 400}
                appLogger.info("Response to " + str(request.client.host) + ": 400")
                return resp

        appLogger.info("Response to " + str(request.client.host) + " File successfully uploaded")
        resp = {'status': 200, 'message': 'File successfully uploaded', 'Document_Id': Doc_Id}
        return resp
    except:
        e = sys.exc_info()
        err = str(e[1])
        appLogger.exception(str(request.client.host) + "  Exception occurred in the /file-upload request!! " + str(e))
        resp = {'status': 501, 'Exception': err}
        return resp


@app.get("/preprocess/{Doc_Id}")
def classify(request: Request, Doc_Id, classify: str = 'TRUE'):
    try:
        appLogger.info("Received GET request from " + str(request.client.host))
        OutputDir = "imag/"
        reqTempDir = 'tmp/' + Doc_Id
        reqResDir = 'res/' + Doc_Id
        inputFilesDirAWS = os.path.join(DirS3, Doc_Id)
        batchname = ""
        currentDir = os.getcwd()
        if not os.path.exists(os.path.dirname(reqResDir)):
            os.mkdir(reqResDir)
        else:
            os.system("rm -rf " + reqResDir)
            os.mkdir(reqResDir)
        if not os.path.exists(reqTempDir):
            os.mkdir(reqTempDir)
            os.chdir(reqTempDir)

            downloadFileObjAWS(inputFilesDirAWS + "/")
            
            tif_file_in_path = glob("*.tif")
            if len(tif_file_in_path) > 0:
                for tiff_file in tif_file_in_path:
                    removetiff(tiff_file)

            os.chdir(currentDir)
        files = os.listdir(os.path.join(currentDir, reqTempDir))
        OutputDir = OutputDir + str(round(time.time())) + '/'
        if not os.path.exists(OutputDir):
            os.makedirs(OutputDir)
        appLogger.info("")
        resp = {'status': 200, 'Document_Id': Doc_Id}
        file_array = []
        image_junk = False

        for file in files:
            if not os.path.isfile(os.path.join(currentDir,reqTempDir, file)):
                continue
            else:
                try:
                    # os.chdir(currentDir)
                    filetype = os.path.splitext(file)[1]
                    filePath = os.path.join(currentDir,reqTempDir, file)
                    total_words = 0
                    out_text = ""
                    filtered_text = ""
                    isJunk = ""
                    ocr_result = []
                    Ocr_File__text_output = []
                    JunkpagesToRemove = []

                    if filetype == '.pdf' or filetype == '.PDF':
                        # pages = convert_from_path(filePath, dpi=600, fmt='jpg',thread_count=1,strict=False)
                        try:
                            inputpdf = PdfFileReader(open(filePath,"rb"))
                            maxPages = inputpdf.numPages
                        except Exception:
                            inputpdf = Pdf.open(filePath)
                            maxPages = len(inputpdf.pages)

                        if maxPages >1:
                            total = []

                            i = 1
                            for page in range(1, maxPages, 10):
                                pages = convert_from_path(filePath, dpi=600, first_page=page,
                                                                                last_page=min(page + 10 - 1, maxPages), fmt= 'jpg',
                                                                                thread_count=1, userpw=None,
                                                                                use_cropbox=False, strict=False)
                                for image in pages:
                                    myfile = OutputDir + 'output' + str(i) + '.jpg'
                                    image.save(myfile, 'JPEG')

                                    isJunk, total, out_text, ocr_res = getConfidence(myfile, Doc_Id, total)
                                    if isJunk == "True":
                                        JunkpagesToRemove.append(i-1)
                                        filtered_text = filtered_text + out_text
                                    else:
                                        filtered_text = filtered_text + out_text
                                    # NP storing ocr text for similarity check
                                    Ocr_File__text_output.append(out_text)

                                    ocr_result.append(ocr_res)
                                    i = i + 1

                        elif maxPages == 1:
                            pages = convert_from_path(filePath, dpi=600, fmt='jpg')
                            counter = 1
                            total = []
                            for i, page in enumerate(pages):
                                myfile = OutputDir + 'output' + str(counter) + '.jpg'
                                page.save(myfile, "JPEG")

                                isJunk, total, out_text, ocr_res = getConfidence(myfile, Doc_Id, total)
                                if isJunk == "True":
                                    JunkpagesToRemove.append(i)
                                    filtered_text = filtered_text + out_text
                                else:
                                    filtered_text = filtered_text + out_text
                                # NP storing ocr text for similarity check
                                Ocr_File__text_output.append(out_text)

                                ocr_result.append(ocr_res)


                        numPages = maxPages
                        total = [float(num) for num in total]
                        tot = sum(total) / len(total)

                    elif filetype == '.jpeg' or filetype == '.jpg' or filetype == '.JPEG' or filetype == '.JPG':
                        numPages = 1
                        total = []
                        results = []
                        isJunk, total, out_text, ocr_res = getConfidence(filePath, total, Doc_Id)
                        tot = sum(total) / len(total)
                        if isJunk == "True":
                            image_junk = True
                    elif filetype == '.docx' or filetype == '.DOCX':
                        tot = 99
                        page_count = 1
                        out_text = readDocx(filePath)
                        numPages = docx_get_characters_number(filePath)

                    elif filetype == '.txt' or filetype == '.TXT':
                        tot = 99
                        page_count = 1
                        f = open(filePath, "r")
                        out_text = f.read()
                        numPages = 1

                    elif filetype == '.xlsx' or filetype == '.XLSX':
                        tot = 99
                        page_count = 1
                        data = readXlsx(filePath)
                        for i in range(0, len(data)):
                            out_text = out_text + " " + data[i]
                        numPages = len(data)

                    if out_text.strip() == "":
                        resp = {'status': 501, 'message': 'No text read from file'}
                        return resp
                    result = {}
                    if (tot < minAverageQuality):
                        result.update({"File_Name": file, "Quality": "Poor", "Page_Count": numPages,
                                       "Quality_Confidence_Score": round(tot, 2)})
                        appLogger.info(
                            "Result of the quality check request from " + str(request.client.host) + ": " + str(result))

                    elif (tot >= minAverageQuality and tot < minGoodQuality):
                        result.update({"File_Name": file, "Quality": "Average", "Page_Count": numPages,
                                       "Quality_Confidence_Score": round(tot, 2)})
                        appLogger.info(
                            "Result of the quality check request from " + str(request.client.host) + ": " + str(result))
                    else:
                        result.update({"File_Name": file, "Quality": "Good", "Page_Count": numPages,
                                       "Quality_Confidence_Score": round(tot, 2)})
                        appLogger.info(
                            "Result of the quality check request from " + str(request.client.host) + ": " + str(result))
                    # file_array.append(result)

                    if (tot < minAverageQuality) or classify.lower() == "false" or len(filtered_text) < 1 or image_junk == True:
                        result.update({'Document_Type': "UNK", 'Classification_Confidence_Score': "N/A",
                                       'Classification_Confidence_Percent': "N/A"})
                        file_array.append(result)
                        resp["File_Names_List"] = file_array
                        # shutil.rmtree("imag/")
                        os.system("rm -rf " + reqResDir)
                        os.system("rm -rf " + reqTempDir)
                        appLogger.info(
                            "Result for the classify request from " + str(request.client.host) + ": " + str(resp))
                        appLogger.info("Returned result successfully to " + str(request.client.host))
                        return (resp)

                    docBatchID = insertBatchDetail(file, numPages, Doc_Id)
                    keyMaster = getData('Key_Master')
                    keyAlias = getData('Key_Alias')
                    keyMapping = getData('Document_Key_Mapping')
                    docMaster = getData('Document_Master')
                    # Look up term - Business Logics from database
                    keyfound = []
                    for row in range(0, len(keyMaster)):
                        total_words += 1
                        occurences = 0
                        keyList = [str(keyMaster[row][1]).lower()]

                        for word in range(0, len(keyAlias)):
                            if keyAlias[word][1] == keyMaster[row][0]:
                                keyList.append(str(keyAlias[word][2]))

                        for key in range(0, len(keyList)):
                            occurences = occurences + len(re.findall(str(keyList[key]).lower(), filtered_text.lower()))
                            if occurences > 0:
                                keyfound.append(keyList[key])
                                insertBatchProcess(docBatchID, keyMaster[row][0], occurences)
                                break

                    docType, confidenceVal, percent = getDocType(docBatchID, keyMapping, docMaster)
                    DocTypeMaster = getData('Document_Types')
                    for DocType in DocTypeMaster:
                        if DocType[1].lower() == docType.lower():
                            docType = DocType[0]
                            result.update({'Document_Type': "INV", 'Classification_Confidence_Score': confidenceVal,
                                           'Classification_Confidence_Percent': percent})
                            break
                    
                    appLogger.info("Junk Pages to Remove" + str(JunkpagesToRemove))
                    # pages_to_delete = IdentifyDuplicatePages(Ocr_File__text_output)
                    pages_to_delete = []
                    for i in JunkpagesToRemove:
                        pages_to_delete.append(i)
                    #
                    #     # appending junk pages to remove list
                    pages_to_delete_set = set(pages_to_delete)
                    appLogger.info("pages_to_delete_set" + str(pages_to_delete_set))
                    # pages_to_delete = [3, 4, 5]  # page numbering starts from 0

                    try:
                        infile = PdfFileReader(filePath,strict=False)

                        newFileName = os.path.splitext(file)[0] + '_edited.pdf'
                        newfilesavedpath = os.path.join(reqTempDir,"Edited", newFileName)

                        if not os.path.exists(reqTempDir+"/Edited"):
                            os.mkdir(reqTempDir+"/Edited")


                        output = PdfFileWriter()
                        editedPageCount=0
                        for i in range(infile.getNumPages()):
                            if i not in pages_to_delete_set:
                                p = infile.getPage(i)
                                editedPageCount+=1
                                output.addPage(p)


                        with open(reqTempDir + '/Edited/' + newFileName, 'wb') as f:
                            output.write(f)
                    
                    except Exception:
                        infile = Pdf.open(filePath)

                        newFileName = os.path.splitext(file)[0] + '_edited.pdf'
                        newfilesavedpath = os.path.join(reqTempDir,"Edited", newFileName)

                        if not os.path.exists(reqTempDir+"/Edited"):
                            os.mkdir(reqTempDir+"/Edited")


                        output = Pdf.new()
                        editedPageCount=0
                        for n, page in enumerate(infile.pages):
                            editedPageCount+=1
                            if n not in pages_to_delete_set:
                                output.pages.append(page)

                        output.save(newfilesavedpath)

                    uploadFileAWS(newfilesavedpath, inputFilesDirAWS + "/Edited/"+ newFileName)
                    fullPathAWS = "https://" + bucketName + ".s3.amazonaws.com/" + inputFilesDirAWS + "/Edited/" + quote(
                        newFileName)


                    updateDocLocation(Doc_Id, editedPageCount,fullPathAWS , True)
                    resp["UpdatedFileLocation"] = fullPathAWS

                    vendorname = IdentifyVendor(Ocr_File__text_output[0][:170])
                    print("vendorname:",vendorname)


                    extractedResultList = []
                    vendors=GetVendorList()

                    #add a function will chek if we need to extract vendor from configured DB
                    if vendorname in vendors:
                        result['vendorname'] = vendorname

                        targetSystempreference = GetOCRPreference(vendorname)


    #Flag to add override the
                        if PreprocessConfig.EnableExtraction and (targetSystempreference=='I2D' or targetSystempreference=='QBOCR'):
                            if targetSystempreference=='I2D':
                                # print(vendorname,"line 1181")
                                extractedResult=ExtractInvoice2Data(filePath,vendorname)
                            elif targetSystempreference =='QBOCR':
                                extractedResult = ExtractData(ocr_result[0], vendorname)
                            isExtractionValid=ValidateOCRresults(extractedResult)
                            if isExtractionValid:
                               batchname=InsertExractedResults(extractedResult,vendorname,Doc_Id,numPages,'QBOCR')

                            extractedResultList.append(extractedResult)
                            extractedResultList.append(isExtractionValid)



                        #result['extractedResult'] = extractedResultList
                            result['batchname'] = batchname

                    file_array.append(result)

                    #---Remove all temp data--
                    if os.path.exists(reqResDir):
                    #    shutil.rmtree("imag/")
                        os.system("rm -rf " + reqResDir)
                        os.system("rm -rf " + reqTempDir)


                except:
                    e = sys.exc_info()
                    traceback.print_exc()
                    if os.path.exists(reqResDir):
                        shutil.rmtree("imag/")
                        os.system("rm -rf " + reqResDir)
                        os.system("rm -rf " + reqTempDir)
                    resp_temp = {'message': 'Error occured in document ' + file}
                    file_array.append(resp_temp)
                    appLogger.error("Exception occured in document  & Error -" + str(e))

            resp["File_Names_List"] = file_array


            appLogger.info("Result for the classify request from " + str(request.client.host) + ": " + str(resp))
            appLogger.info("Returned result successfully to " + str(request.client.host))
            return (resp)
    except:
            traceback.print_exc()
            e = sys.exc_info()
            if os.path.exists(reqResDir):
                shutil.rmtree("imag/")
                os.system("rm -rf " + reqResDir)
                os.system("rm -rf " + reqTempDir)
            resp_temp = {'message': 'Error occured while processing document'}
            appLogger.error("Exception occured in processing -" + str(e))
            return (resp_temp)

def getClasResult(image_path):
    url = 'https://app.nanonets.com/api/v2/ImageCategorization/LabelFile/'

    data = {'file': open(f'{image_path}', 'rb'), 'modelId': ('', 'ec920477-bb29-4645-babe-250346d037fc')}

    response = requests.post(url, auth= requests.auth.HTTPBasicAuth('K20t_QuKhhv4RQyzasqzBW68HSkwsffI', ''), files=data)

    result_dict = response.json()
    print("result_dict:",result_dict['result'][0]['prediction'])
    
    return result_dict

def getClassification(myfile, Doc_Id):
    try:
        result_dict = getClasResult(myfile)
        if result_dict['result'][0]['prediction'][0]['label'] == 'INV':
            isJunk = False
        elif result_dict['result'][0]['prediction'][0]['label'] == 'JUNK':
            isJunk = True
        print("isJunk:",isJunk)
        return isJunk,result_dict
    
    except:
        e = sys.exc_info()
        traceback.print_exc()
        appLogger.exception("Error occured while uploading file: " + str(e))
        return False

@app.get("/classify_nano/{Doc_Id}")
def classify(request: Request, Doc_Id , classify: str = 'TRUE'):
    try:
        appLogger.info("Received GET request from " + str(request.client.host))
        OutputDir = "imag/"
        reqTempDir = 'tmp/' + Doc_Id
        reqResDir = 'res/' + Doc_Id
        inputFilesDirAWS = os.path.join(DirS3, Doc_Id)
        batchname = ""
        currentDir = os.getcwd()
        if not os.path.exists(os.path.dirname(reqResDir)):
            os.mkdir(reqResDir)
        else:
            os.system("rm -rf " + reqResDir)
            os.mkdir(reqResDir)
        if not os.path.exists(reqTempDir):
            os.mkdir(reqTempDir)
            os.chdir(reqTempDir)
            downloadFileObjAWS(inputFilesDirAWS + "/")
            os.chdir(currentDir)
        files = os.listdir(os.path.join(currentDir, reqTempDir))
        OutputDir = OutputDir + str(round(time.time())) + '/'
        if not os.path.exists(OutputDir):
            os.makedirs(OutputDir)
        appLogger.info("")
        resp = {'status': 200, 'Document_Id': Doc_Id}
        file_array = []
        image_junk = False

        for file in files:
            if not os.path.isfile(os.path.join(currentDir,reqTempDir, file)):
                continue
            else:
                try:
                    # os.chdir(currentDir)
                    filetype = os.path.splitext(file)[1]
                    filePath = os.path.join(currentDir,reqTempDir, file)
                    JunkpagesToRemove = []

                    if filetype == '.pdf' or filetype == '.PDF':
                        
                        pages = convert_from_path(filePath, dpi=600, fmt='jpg')
                        counter = 1
                        docstype = []
                        conf_per = {}
                        for i, page in enumerate(pages):
                        
                            myfile = OutputDir + 'output' + str(counter) + '.jpg'
                            page.save(myfile, "JPEG")

                            isJunk,result_dict = getClassification(myfile,Doc_Id)

                            if isJunk == True:
                                JunkpagesToRemove.append(i)
                                docstype.append('JUNK')
                            else:
                                docstype.append('INV')
                        
                        numPages = len(pages)
                        if 'INV' in docstype:
                            docType = 'INV'
                        else:
                            docType = 'UNK' 
                        # print('JunkPages:',JunkpagesToRemove)
                    else:
                        return {'UnsupportedFormatError:':"Please Provide a pdf document"}

                    result = {}

                    result.update({"File_Name": file, "Page_Count": numPages})

                    result.update({'Document_Type': docType})

                    file_array.append(result)
                    resp["File_Names_List"] = file_array
                    
                    docBatchID = insertBatchDetail(file, numPages, Doc_Id)

                    # appending junk pages to remove list
                    pages_to_delete_set = set(JunkpagesToRemove)
                    print("pages_to_delete_set:",pages_to_delete_set)

                    # pages_to_delete = [3, 4, 5]  # page numbering starts from 0
                    infile = PdfFileReader(filePath, 'rb')

                    newFileName = os.path.splitext(file)[0] + '_edited.pdf'
                    newfilesavedpath = os.path.join(reqTempDir,"Edited", newFileName)
                    output = PdfFileWriter()
                    if not os.path.exists(reqTempDir+"/Edited"):
                        os.mkdir(reqTempDir+"/Edited")


                    editedPageCount=0
                    for i in range(infile.getNumPages()):
                        if i not in pages_to_delete_set:
                            p = infile.getPage(i)
                            editedPageCount+=1
                            output.addPage(p)

                    with open(reqTempDir + '/Edited/' + newFileName, 'wb') as f:
                        output.write(f)

                    uploadFileAWS(newfilesavedpath, inputFilesDirAWS + "/Edited/"+ newFileName)
                    fullPathAWS = "https://" + bucketName + ".s3.amazonaws.com/" + inputFilesDirAWS + "/Edited/" + quote(
                        newFileName)

                    updateDocLocation(Doc_Id, editedPageCount,fullPathAWS , True)
                    resp["UpdatedFileLocation"] = fullPathAWS

                    #---Remove all temp data--
                    # if os.path.exists(reqResDir):
                    # #    shutil.rmtree("imag/")
                    #    shutil.rmtree(reqTempDir)
                    #    shutil.rmtree(reqResDir)

                except:
                    e = sys.exc_info()
                    traceback.print_exc()
                    if os.path.exists(reqResDir):
                        shutil.rmtree("imag/")
                        os.system("rm -rf " + reqResDir)
                        os.system("rm -rf " + reqTempDir)
                    resp_temp = {'message': 'Error occured in document ' + file}
                    file_array.append(resp_temp)
                    appLogger.error("Exception occured in document  & Error -" + str(e))

            resp["File_Names_List"] = file_array


            appLogger.info("Result for the classify request from " + str(request.client.host) + ": " + str(resp))
            appLogger.info("Returned result successfully to " + str(request.client.host))
            return (resp)
    except:
            traceback.print_exc()
            e = sys.exc_info()
            if os.path.exists(reqResDir):
                shutil.rmtree("imag/")
                os.system("rm -rf " + reqResDir)
                os.system("rm -rf " + reqTempDir)
            resp_temp = {'message': 'Error occured while processing document'}
            appLogger.error("Exception occured in processing -" + str(e))
            return (resp_temp)
    finally:
        del(Doc_Batch_Process[:])
        del(Doc_Batch_Process_tt[:])

if __name__ == '__main__':
    uvicorn.run('NSDIIP:app', host='0.0.0.0', port=8080,reload=True)
